"""
Routes for the RAG-based assessment system
"""
from flask import Blueprint, request, jsonify, send_file
from rag_assessor import get_assessor
from rag_trainer import RAGTrainer
from train_from_csv import main as train_from_csv_main
import os
import csv
import io
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import colorsys
import hashlib

rag_routes = Blueprint('rag_routes', __name__)

def format_assessment_result(result, student_name, repo_url):
    """Format assessment result to ensure consistent structure"""
    if 'error' in result:
        return {
            "student": {
                "name": student_name,
                "github_url": repo_url
            },
            "status": "error",
            "error": result['error']
        }
    
    return {
        "student": {
            "name": student_name,
            "github_username": result.get("student", {}).get("github_username", ""),
            "repository_name": result.get("student", {}).get("repository_name", ""),
            "repository_url": repo_url
        },
        "status": "success",
        "assessment": {
            "criterion": result.get("criterion", ""),
            "mark": result.get("mark", ""),
            "level": result.get("level", ""),
            "justification": result.get("justification", ""),
            "observations": result.get("observations", [])
        }
    }

@rag_routes.route("/rag/assess", methods=["POST"])
def rag_assess():
    """Assess code using the RAG-based system"""
    # Support both form data and JSON
    if request.is_json:
        data = request.json
        code = data.get('code')
        rubric = data.get('rubric')
        repo_url = data.get('repo_url')
    else:
        code = request.form.get('code')
        repo_url = request.form.get('repo_url')
        rubric_file = request.files.get('rubric')
        students_file = request.files.get('students')
        
        # Handle rubric file if provided
        if rubric_file:
            rubric = rubric_file.read().decode('utf-8')
        else:
            rubric = None
            
        # Handle CSV file with multiple students
        if students_file and not code and not repo_url:
            return assess_multiple_students(students_file, rubric)
    
    # If repo_url is provided but no code, we need to fetch the code
    if repo_url and not code:
        from repo_analyzer import analyze_github_repo
        try:
            code = analyze_github_repo(repo_url)
        except Exception as e:
            return jsonify({"error": f"Failed to analyze repository: {str(e)}"}), 500
    
    if not code:
        return jsonify({"error": "No code provided"}), 400
    
    if not rubric:
        return jsonify({"error": "No rubric provided"}), 400
    
    try:
        # Get the RAG assessor
        assessor = get_assessor()
        
        # Assess the code
        result = assessor.assess_code(code, rubric)
        
        return jsonify({"success": True, "result": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def assess_multiple_students(students_file, rubric):
    """Helper function to assess multiple students from a CSV file"""
    try:
        # Read students CSV
        students_csv = students_file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(students_csv))
        
        # Get the RAG assessor
        assessor = get_assessor()
        
        # Assess each student's repository
        results = []
        for row in csv_reader:
            student_name = row.get('name', '').strip()
            repo_url = row.get('repo_url', '').strip()
            
            if not student_name or not repo_url:
                continue
                
            try:
                # Analyze repository
                from repo_analyzer import analyze_github_repo
                code = analyze_github_repo(repo_url)
                
                # Assess the code
                assessment = assessor.assess_code(code, rubric)
                results.append(format_assessment_result(assessment, student_name, repo_url))
            except Exception as e:
                results.append(format_assessment_result(
                    {"error": str(e)}, student_name, repo_url
                ))
        
        return jsonify({
            "success": True, 
            "results": results,
            "summary": {
                "total_students": len(results),
                "successful_assessments": sum(1 for r in results if r.get("status") == "success"),
                "failed_assessments": sum(1 for r in results if r.get("status") == "error")
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@rag_routes.route("/rag/train/csv", methods=["POST"])
def train_from_csv():
    """Train the RAG system from a CSV file"""
    if 'file' not in request.files:
        return jsonify({"error": "No CSV file provided"}), 400
        
    file = request.files['file']
    if not file.filename.endswith('.csv'):
        return jsonify({"error": "File must be a CSV"}), 400
    
    try:
        # Save the uploaded CSV temporarily
        temp_path = 'temp_training.csv'
        file.save(temp_path)
        
        # Run the training script
        train_from_csv_main(temp_path)
        
        # Clean up
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        return jsonify({
            "success": True,
            "message": "RAG system trained successfully from CSV"
        })
    except Exception as e:
        # Clean up on error
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"error": str(e)}), 500

@rag_routes.route("/rag/train", methods=["POST"])
def train_from_drive():
    """Train the RAG system from Google Drive folders"""
    data = request.json
    
    rubrics_folder_id = data.get('rubrics_folder_id') or os.getenv("RUBRICS_FOLDER_ID")
    examples_folder_id = data.get('examples_folder_id') or os.getenv("EXAMPLES_FOLDER_ID")
    
    if not rubrics_folder_id:
        return jsonify({"error": "No rubrics folder ID provided"}), 400
    
    if not examples_folder_id:
        return jsonify({"error": "No examples folder ID provided"}), 400
    
    try:
        trainer = RAGTrainer()
        trainer.bulk_train_from_drive(rubrics_folder_id, examples_folder_id)
        
        return jsonify({
            "success": True, 
            "message": "RAG system trained successfully"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@rag_routes.route("/rag/train/manual", methods=["POST"])
def train_manual():
    """Manually add a rubric and example to the RAG system"""
    data = request.json
    
    if not data:
        return jsonify({"error": "No data provided"}), 400
    
    try:
        trainer = RAGTrainer()
        
        # Add rubric if provided
        if 'rubric' in data:
            rubric_text = data['rubric']
            rubric_id = data.get('rubric_id')
            trainer.add_rubric(rubric_text, rubric_id)
        
        # Add example if provided
        if 'example' in data and isinstance(data['example'], dict):
            example = data['example']
            if 'code' in example and 'scores' in example:
                trainer.add_assessment_example(
                    example['code'],
                    data.get('rubric', ''),  # Use the rubric if provided
                    example['scores'],
                    example.get('id')
                )
        
        return jsonify({
            "success": True, 
            "message": "Training data added successfully"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
@rag_routes.route("/rag/assess/excel", methods=["POST"])
def rag_assess_excel():
    """Assess student code repositories using RAG-based system and return results as Excel file"""
    students_file = request.files.get('students')
    rubric_file = request.files.get('rubric')
    
    if not students_file:
        return jsonify({"error": "No students CSV file provided"}), 400
    
    if not rubric_file:
        return jsonify({"error": "No rubric file provided"}), 400
    
    try:
        # Read rubric
        rubric_content = rubric_file.read().decode('utf-8')
        
        # Parse rubric to extract all criteria and scoring levels
        rubric_data = parse_rubric(rubric_content)
        
        # Read students CSV
        students_csv = students_file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(students_csv))
        
        # Get the RAG assessor
        assessor = get_assessor()
        
        # Assess each student's repository
        results = []
        for row in csv_reader:
            student_name = row.get('name', '').strip()
            repo_url = row.get('repo_url', '').strip()
            
            if not student_name or not repo_url:
                continue
                
            try:
                # Assess the code from GitHub
                assessment = assessor.assess_code(repo_url, rubric_content)
                results.append(format_assessment_result(assessment, student_name, repo_url))
            except Exception as e:
                results.append(format_assessment_result(
                    {"error": str(e)}, student_name, repo_url
                ))
        
        # Create Excel file from results
        excel_data = create_excel_from_results(results, rubric_data)
        
        # Return Excel file
        return send_file(
            excel_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='assessment_results.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def parse_rubric(rubric_content):
    """Parse rubric to extract criteria and scoring levels"""
    rubric_data = {
        'criteria': [],
        'scoring_levels': {}
    }
    
    current_criterion = None
    
    for line in rubric_content.split('\n'):
        line = line.strip()
        if not line:
            continue
            
        if line.startswith('Main Criterion:'):
            current_criterion = line.split(':', 1)[1].strip()
            rubric_data['criteria'].append(current_criterion)
            rubric_data['scoring_levels'][current_criterion] = {}
        elif '(No Mark):' in line and current_criterion:
            mark = '0'
            description = line.split(':', 1)[1].strip()
            rubric_data['scoring_levels'][current_criterion][mark] = description
        elif 'Marks:' in line and current_criterion:
            mark = line.split('Marks:', 1)[0].strip()
            description = line.split(':', 1)[1].strip()
            rubric_data['scoring_levels'][current_criterion][mark] = description
    
    return rubric_data

def generate_color_for_criterion(criterion_name):
    """Generate a consistent color for a criterion based on its name"""
    # Predefined distinct colors for better visual separation
    distinct_colors = [
        "4A90E2",  # Blue
        "50E3C2",  # Teal
        "F5A623",  # Orange
        "7ED321",  # Green
        "B8E986",  # Light Green
        "BD10E0",  # Purple
        "D0021B",  # Red
        "9013FE",  # Violet
        "417505",  # Dark Green
        "8B572A",  # Brown
        "F8E71C",  # Yellow
        "9B9B9B",  # Gray
        "4A4A4A",  # Dark Gray
        "FF5E3A",  # Coral
        "FF9500",  # Amber
        "FFCC00",  # Gold
    ]
    
    # Create a hash of the criterion name for consistency
    hash_obj = hashlib.md5(criterion_name.encode())
    hash_int = int(hash_obj.hexdigest(), 16)
    
    # Use the hash to select a color from the predefined list
    color_index = hash_int % len(distinct_colors)
    
    return distinct_colors[color_index]

def create_excel_from_results(results, rubric_data):
    """Create beautifully formatted Excel file from assessment results with student-based rows"""
    # Create a Pandas Excel writer
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Generate dynamic colors for criteria
    criterion_colors = {}
    for criterion in rubric_data['criteria']:
        criterion_colors[criterion] = generate_color_for_criterion(criterion)

    # Create a student-based assessment sheet
    # First, prepare the data structure
    students = []
    for result in results:
        students.append({
            'name': result['student']['name'],
            'github_username': result['student'].get('github_username', ''),
            'repository_name': result['student'].get('repository_name', ''),
            'repository_url': result['student'].get('repository_url', ''),
            'status': result['status'],
            'assessment': result.get('assessment', {}) if result['status'] == 'success' else {}
        })
    
    # Create the DataFrame with students as rows and criteria as columns
    # First column will be student names, followed by basic info, then criteria
    assessment_data = {
        'Student Name': [],
        'GitHub Username': [],
        'Repository': [],
        'Repository URL': [],
        'Status': []
    }
    
    # Add columns for each criterion
    for criterion in rubric_data['criteria']:
        assessment_data[f"{criterion} - Mark"] = []
        assessment_data[f"{criterion} - Level"] = []
    
    # Add rows for each student
    for student in students:
        assessment_data['Student Name'].append(student['name'])
        assessment_data['GitHub Username'].append(student['github_username'])
        assessment_data['Repository'].append(student['repository_name'])
        assessment_data['Repository URL'].append(student['repository_url'])
        assessment_data['Status'].append(student['status'])
        
        # Add assessment data for each criterion
        for criterion in rubric_data['criteria']:
            if student['status'] == 'success' and student['assessment'].get('criterion') == criterion:
                assessment_data[f"{criterion} - Mark"].append(student['assessment'].get('mark', 'N/A'))
                assessment_data[f"{criterion} - Level"].append(student['assessment'].get('level', 'N/A'))
            else:
                assessment_data[f"{criterion} - Mark"].append('N/A')
                assessment_data[f"{criterion} - Level"].append('N/A')
    
    # Create the assessment DataFrame
    assessment_df = pd.DataFrame(assessment_data)
    assessment_df.to_excel(writer, sheet_name='Assessment', index=False)
    
    # Style the assessment sheet
    assessment_sheet = writer.sheets['Assessment']
    style_student_assessment_sheet(assessment_sheet, assessment_df, rubric_data['criteria'], criterion_colors)
    
    # Create detailed sheets for each student
    for i, student in enumerate(students):
        if student['status'] == 'success':
            student_name = student['name']
            sheet_name = f"{student_name[:28]}" if len(student_name) > 28 else student_name
            
            # Handle duplicate sheet names
            if sheet_name in writer.sheets:
                sheet_name = f"{sheet_name[:25]}_{i}"
            
            # Create student data
            student_data = {
                'Information': ['Name', 'GitHub Username', 'Repository Name', 'Repository URL', '', 'Assessment Results'],
                'Value': [
                    student_name,
                    student['github_username'],
                    student['repository_name'],
                    student['repository_url'],
                    '',
                    ''
                ]
            }
            
            # Add assessment data
            if 'assessment' in student and student['assessment']:
                criterion = student['assessment'].get('criterion', '')
                student_data['Information'].extend(['', criterion])
                student_data['Value'].extend(['', ''])
                
                student_data['Information'].extend(['Mark', 'Level'])
                student_data['Value'].extend([
                    student['assessment'].get('mark', 'N/A'),
                    student['assessment'].get('level', 'N/A')
                ])
                
                # Add justification
                student_data['Information'].extend(['', 'Justification'])
                student_data['Value'].extend(['', student['assessment'].get('justification', '')])
                
                # Add observations
                student_data['Information'].append('')
                student_data['Value'].append('')
                student_data['Information'].append('Observations')
                student_data['Value'].append('')
                
                for j, observation in enumerate(student['assessment'].get('observations', [])):
                    student_data['Information'].append(f'Observation {j+1}')
                    student_data['Value'].append(observation)
            
            # Create student DataFrame and write to Excel
            student_df = pd.DataFrame(student_data)
            student_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Style the student sheet
            student_sheet = writer.sheets[sheet_name]
            style_student_sheet(student_sheet, student_df, rubric_data['criteria'], criterion_colors)

    # Add rubric sheet
    rubric_data_for_excel = {
        'Criterion': [],
        'Mark': [],
        'Level Description': []
    }
    
    for criterion in rubric_data['criteria']:
        for mark, description in rubric_data['scoring_levels'][criterion].items():
            rubric_data_for_excel['Criterion'].append(criterion)
            rubric_data_for_excel['Mark'].append(mark)
            rubric_data_for_excel['Level Description'].append(description)
    
    rubric_df = pd.DataFrame(rubric_data_for_excel)
    rubric_df.to_excel(writer, sheet_name='Rubric', index=False)
    
    # Style the rubric sheet
    rubric_sheet = writer.sheets['Rubric']
    style_rubric_sheet(rubric_sheet, rubric_df, criterion_colors)

    # Save the Excel file to the BytesIO object
    writer.close()
    output.seek(0)

    return output

def style_summary_sheet(worksheet, df, criteria, criterion_colors):
    """Apply beautiful styling to the summary sheet"""
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Style basic info headers (Student Name, GitHub Username, etc.)
    basic_info_fill = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')
    
    # Get the number of basic columns (before criteria columns)
    basic_columns = 8  # Student Name, GitHub Username, Repository Name, Repository URL, Status, Average Score, Percentage, Pass/Fail
    
    # Style basic info headers
    for col in range(1, basic_columns + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = basic_info_fill
        cell.border = thin_border
        cell.alignment = wrap_alignment
        # Set column width
        worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Style criterion headers with their respective colors
    col_index = basic_columns + 1
    for criterion in criteria:
        color = criterion_colors.get(criterion, 'CCCCCC')
        criterion_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Style both Mark and Level columns for this criterion
        for suffix in [' - Mark', ' - Level']:
            cell = worksheet.cell(row=1, column=col_index)
            cell.font = header_font
            cell.fill = criterion_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment
            # Set column width
            worksheet.column_dimensions[get_column_letter(col_index)].width = 12
            col_index += 1
    
    # Style data rows
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Add alternating row colors for better readability
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 25

def style_student_sheet(worksheet, df, criteria, criterion_colors):
    """Apply beautiful styling to student detail sheets"""
    
    # Define styles
    header_font = Font(bold=True, color='000000', size=12)
    info_font = Font(bold=True, color='000000', size=11)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    left_alignment = Alignment(horizontal='left', vertical='top')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Style the sheet
    for row in range(1, len(df) + 1):
        info_cell = worksheet.cell(row=row, column=1)  # Information column
        value_cell = worksheet.cell(row=row, column=2)  # Value column
        
        info_cell.border = thin_border
        value_cell.border = thin_border
        info_cell.alignment = left_alignment
        value_cell.alignment = wrap_alignment
        
        # Style based on content
        if info_cell.value in ['Name', 'GitHub Username', 'Repository Name', 'Repository URL']:
            info_cell.font = info_font
            info_cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        elif info_cell.value == 'Assessment Results':
            info_cell.font = header_font
            info_cell.fill = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')
            info_cell.font = Font(bold=True, color='FFFFFF', size=12)
        elif any(criterion in str(info_cell.value) for criterion in criteria):
            # Color code based on criterion
            for criterion in criteria:
                if criterion in str(info_cell.value):
                    color = criterion_colors.get(criterion, 'CCCCCC')
                    info_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    info_cell.font = info_font
                    break
        elif info_cell.value == 'Observations':
            info_cell.font = header_font
            info_cell.fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        else:
            info_cell.font = data_font
            value_cell.font = data_font
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 60

def style_student_assessment_sheet(worksheet, df, criteria, criterion_colors):
    """Apply beautiful styling to the assessment sheet with student-based rows and criteria columns"""
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    student_font = Font(bold=True, size=10)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Get the number of basic info columns
    basic_columns = 5  # Student Name, GitHub Username, Repository, Repository URL, Status
    
    # Style header row with different colors for each criterion
    header_fill = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')
    
    # Style basic info headers
    for col in range(1, basic_columns + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
        # Set column width
        worksheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Style criterion headers with their respective colors
    col_index = basic_columns + 1
    for criterion in criteria:
        color = criterion_colors.get(criterion, 'CCCCCC')
        criterion_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Style both Mark and Level columns for this criterion
        for suffix in [' - Mark', ' - Level']:
            cell = worksheet.cell(row=1, column=col_index)
            cell.font = header_font
            cell.fill = criterion_fill
            cell.border = thin_border
            cell.alignment = center_alignment
            # Set column width
            worksheet.column_dimensions[get_column_letter(col_index)].width = 15
            col_index += 1
    
    # Style data rows (student rows)
    for row in range(2, len(df) + 2):
        # Style student name cell
        student_cell = worksheet.cell(row=row, column=1)
        student_cell.font = student_font
        student_cell.border = thin_border
        student_cell.alignment = center_alignment
        
        # Style basic info cells
        for col in range(1, basic_columns + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = data_font
            
            # Add alternating row colors for better readability
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Style criterion data cells with lighter shades of the criterion colors
        col_index = basic_columns + 1
        for criterion in criteria:
            color = criterion_colors.get(criterion, 'CCCCCC')
            
            # Convert to a lighter shade
            r = int(color[0:2], 16)
            g = int(color[2:4], 16)
            b = int(color[4:6], 16)
            
            # Convert to HSV, adjust lightness, convert back to RGB
            h, s, v = colorsys.rgb_to_hsv(r/255, g/255, b/255)
            v = min(1.0, v * 1.3)  # Increase brightness by 30%
            s = max(0.1, s * 0.7)  # Reduce saturation by 30%
            r, g, b = colorsys.hsv_to_rgb(h, s, v)
            
            light_color = f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"
            light_fill = PatternFill(start_color=light_color, end_color=light_color, fill_type='solid')
            
            # Style both Mark and Level cells for this criterion
            for _ in range(2):  # Mark and Level
                cell = worksheet.cell(row=row, column=col_index)
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = data_font
                cell.fill = light_fill
                col_index += 1
    
    # Set row heights
    worksheet.row_dimensions[1].height = 30  # Header row
    for row in range(2, len(df) + 2):
        worksheet.row_dimensions[row].height = 22

def style_rubric_sheet(worksheet, df, criterion_colors):
    """Apply beautiful styling to the rubric sheet"""
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Style headers
    header_fill = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')
    for col in range(1, 4):  # Criterion, Mark, Level Description
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Style data rows
    current_criterion = None
    for row in range(2, len(df) + 2):
        criterion_cell = worksheet.cell(row=row, column=1)
        mark_cell = worksheet.cell(row=row, column=2)
        description_cell = worksheet.cell(row=row, column=3)
        
        # Apply borders
        for cell in [criterion_cell, mark_cell, description_cell]:
            cell.border = thin_border
            cell.font = data_font
        
        # Color code by criterion
        if criterion_cell.value and criterion_cell.value != current_criterion:
            current_criterion = criterion_cell.value
        
        if current_criterion and current_criterion in criterion_colors:
            color = criterion_colors[current_criterion]
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            criterion_cell.fill = fill
            criterion_cell.font = Font(bold=True, color='FFFFFF', size=10)
        
        # Set alignments
        criterion_cell.alignment = wrap_alignment
        mark_cell.alignment = center_alignment
        description_cell.alignment = wrap_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 8
    worksheet.column_dimensions['C'].width = 50