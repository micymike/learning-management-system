# Multi-Agent CSV Processing System
import asyncio
from agents.orchestrator import AgentOrchestrator
from agents.csv_agent import CSVAgent
import csv
from io import StringIO
from datetime import datetime
import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
from flask import send_file
import io
import json

class ScoreManager:
    def __init__(self):
        self.scores = []
        self.next_id = 1

    def add_scores(self, new_scores):
        """Add scores with unique IDs"""
        for score in new_scores:
            score['id'] = self.next_id
            self.next_id += 1
            self.scores.append(score)
        return self.scores

    def delete_score(self, score_id):
        """Delete a single score by ID"""
        self.scores = [score for score in self.scores if score['id'] != score_id]
        return self.scores

    def delete_scores(self, score_ids):
        """Delete multiple scores by IDs"""
        self.scores = [score for score in self.scores if score['id'] not in score_ids]
        return self.scores

    def get_scores(self):
        """Get all scores"""
        return self.scores

    def clear_scores(self):
        """Clear all scores"""
        self.scores = []
        return self.scores

# Create a global instance of ScoreManager
score_manager = ScoreManager()

def process_csv(file_storage):
    """
    Accepts a Flask file storage object, parses CSV, Excel, or TXT, and returns a list of dicts with relevant student information.
    Supports multiple formats by detecting the file extension.
    """
    filename = file_storage.filename.lower()
    results = []
    print(f"Processing file: {filename}")  # Debug log

    if filename.endswith('.csv'):
        content = file_storage.read().decode('utf-8')
        print(f"CSV Content: {content}")  # Debug log
        file_storage.seek(0)
        reader = csv.reader(StringIO(content))
        print(f"Headers: {next(reader, [])}")  # Debug log
        file_storage.seek(0)
        reader = csv.reader(StringIO(content))
        headers = next(reader, [])
        headers_lower = [h.lower() for h in headers]
        for row in reader:
            if not any(row):
                continue
            row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            student_data = {}
            for header, value in row_dict.items():
                if 'name' in header.lower():
                    student_data['name'] = value.strip()
                elif any(x in header.lower() for x in ['github', 'repo']) and 'url' in header.lower():
                    student_data['repo_url'] = value.strip()
                elif 'group' in header.lower():
                    student_data['group'] = value.strip()
                elif 'project' in header.lower():
                    student_data['project_name'] = value.strip()
                elif 'deployed' in header.lower() and 'url' in header.lower():
                    student_data['deployed_url'] = value.strip()
            if 'name' in student_data and 'repo_url' in student_data:
                results.append(student_data)
        return results

    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        df = pd.read_excel(file_storage)
        for _, row in df.iterrows():
            student_data = {}
            for col in df.columns:
                col_lower = col.lower()
                value = str(row[col]) if not pd.isnull(row[col]) else ''
                if 'name' in col_lower:
                    student_data['name'] = value.strip()
                elif 'github' in col_lower and 'url' in col_lower:
                    student_data['repo_url'] = value.strip()
                elif 'group' in col_lower:
                    student_data['group'] = value.strip()
                elif 'project' in col_lower:
                    student_data['project_name'] = value.strip()
                elif 'deployed' in col_lower and 'url' in col_lower:
                    student_data['deployed_url'] = value.strip()
            if 'name' in student_data and 'repo_url' in student_data:
                results.append(student_data)
        return results

    elif filename.endswith('.txt'):
        content = file_storage.read().decode('utf-8')
        file_storage.seek(0)
        reader = csv.reader(StringIO(content), delimiter='\t')
        try:
            headers = next(reader, [])
            if len(headers) < 2:
                file_storage.seek(0)
                reader = csv.reader(StringIO(content), delimiter=',')
                headers = next(reader, [])
        except Exception:
            return []
        headers_lower = [h.lower() for h in headers]
        for row in reader:
            if not any(row):
                continue
            row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            student_data = {}
            for header, value in row_dict.items():
                if 'name' in header.lower():
                    student_data['name'] = value.strip()
                elif 'github' in header.lower() and 'url' in header.lower():
                    student_data['repo_url'] = value.strip()
                elif 'group' in header.lower():
                    student_data['group'] = value.strip()
                elif 'project' in header.lower():
                    student_data['project_name'] = value.strip()
                elif 'deployed' in header.lower() and 'url' in header.lower():
                    student_data['deployed_url'] = value.strip()
            if 'name' in student_data and 'repo_url' in student_data:
                results.append(student_data)
        return results

    else:
        raise ValueError('Unsupported file type. Please upload a .csv, .xlsx, .xls, or .txt file.')

def generate_scores_excel(scores):
    """
    Generate an Excel file with student scores in a formatted table
    """
    df = pd.DataFrame(scores)
    if 'id' in df.columns:
        df = df.drop('id', axis=1)  # Remove ID column from Excel export
    
    # Create Excel writer object
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # Write DataFrame to Excel
    df.to_excel(writer, sheet_name='Scores', index=False)
    
    # Access the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Scores']
    
    # Format header
    for col in range(len(df.columns)):
        cell = worksheet.cell(row=1, column=col + 1)
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Auto-adjust column width
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    writer.close()
    output.seek(0)
    return output

def format_scores_for_display(scores):
    """
    Format scores data for display in the web interface
    """
    df = pd.DataFrame(scores)
    return df.to_html(classes=['table', 'table-striped', 'table-bordered'], index=False)

# Multi-Agent System Integration
orchestrator = AgentOrchestrator()

async def process_with_agents(file_storage, rubric):
    """
    Process CSV file and rubric using the multi-agent system
    """
    try:
        filename = file_storage.filename.lower()
        
        if filename.endswith('.csv'):
            content = file_storage.read().decode('utf-8')
            file_storage.seek(0)
            csv_data = {
                'file_content': content,
                'filename': filename
            }
        elif filename.endswith(('.xlsx', '.xls')):
            csv_data = {
                'file_storage': file_storage,
                'filename': filename
            }
        else:
            raise ValueError('Unsupported file type')
        
        # Process using orchestrator
        result = await orchestrator.process_assessment(csv_data, rubric)
        return result
        
    except Exception as e:
        return {'error': str(e)}

def run_agent_processing(file_storage, rubric):
    """
    Synchronous wrapper for async agent processing
    """
    return asyncio.run(process_with_agents(file_storage, rubric))
