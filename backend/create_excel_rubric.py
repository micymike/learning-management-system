import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Rubric"

# Add headers
ws['A1'] = "Criterion"
ws['B1'] = "Max Points"
ws['C1'] = "Level 0"
ws['D1'] = "Level 1-3"
ws['E1'] = "Level 4-8"
ws['F1'] = "Level 10-12"

# Make headers bold
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# Add data
ws['A2'] = ""
ws['B2'] = ""
ws['C2'] = "Incorrect Code - The code provided by the student is fundamentally incorrect, and it does not fulfill the specified requirements."
ws['D2'] = "The code contains some correct elements but also includes significant errors that impact its functionality."
ws['E2'] = "The majority of the code is correct, with only minor errors that do not significantly affect its overall functionality."
ws['F2'] = "The code is entirely correct, meeting all specified requirements and functioning as intended."

# Add more criteria
ws['A3'] = "Code Structure"
ws['B3'] = 8
ws['C3'] = "Poor structure with no clear organization"
ws['D3'] = "Basic structure with some organization"
ws['E3'] = "Well-structured code with clear organization"

ws['A4'] = "Documentation"
ws['B4'] = 8
ws['C4'] = "Little to no documentation"
ws['D4'] = "Basic documentation present"
ws['E4'] = "Comprehensive documentation"

ws['A5'] = "Error Handling"
ws['B5'] = 8
ws['C5'] = "No error handling"
ws['D5'] = "Basic error handling"
ws['E5'] = "Comprehensive error handling"

ws['A6'] = "Testing"
ws['B6'] = 8
ws['C6'] = "No testing"
ws['D6'] = "Basic testing"
ws['E6'] = "Comprehensive testing"

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 40

# Set text wrapping for description cells
for row in range(2, 7):
    for col in range(3, 7):
        if ws.cell(row=row, column=col).value:
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

# Save the workbook
wb.save("working_rubric.xlsx")
print("Excel rubric created: working_rubric.xlsx")
