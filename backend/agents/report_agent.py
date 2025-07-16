from .base_agent import BaseAgent, AgentStatus
from typing import Dict, Any, List
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font

class ReportAgent(BaseAgent):
    def __init__(self):
        super().__init__("report_agent")
        
    async def process(self, data: Dict[str, Any]) -> Dict[str, Any]:
        try:
            results = data.get('results', [])
            rubric = data.get('rubric', None)

            if not results:
                raise ValueError("No results to process")
                
            excel_file = await self._generate_excel_report(results, rubric)
            summary = await self._generate_summary(results)
            
            return {
                # 'excel_file': excel_file,  # Removed to avoid JSON serialization error
                'summary': summary,
                'status': 'completed'
            }
            
        except Exception as e:
            self.status = AgentStatus.ERROR
            return {
                'error': str(e),
                'status': 'error'
            }
    
    async def _generate_excel_report(self, results: List[Dict], rubric: str = None) -> io.BytesIO:
        print("=== DEBUG: Results passed to _generate_excel_report ===")
        import pprint
        pprint.pprint(results)
        print("=== END DEBUG ===")

        # Parse rubric for criterion titles and max points
        rubric_criteria = []
        rubric_map = {}
        if rubric:
            import re
            # Match lines like: "1. Code Structure (3mk): ..." or "1. Code Structure: ... (3mk)"
            for line in rubric.splitlines():
                m = re.match(r"\s*\d+\.\s*([^(:\n]+)[(:].*?(\d+)\s*mk\)?", line, re.IGNORECASE)
                if m:
                    title = m.group(1).strip()
                    points = m.group(2).strip()
                    rubric_criteria.append(f"{title} ({points}mk)")
                    rubric_map[f"Criterion{len(rubric_criteria)}"] = f"{title} ({points}mk)"
        # Fallback: use whatever is in the scores if rubric not provided
        if not rubric_criteria:
            all_criteria = set()
            for result in results:
                scores = result.get('scores', {})
                for criterion in scores.keys():
                    all_criteria.add(criterion)
            rubric_criteria = sorted(all_criteria)
            # Try to map Criterion1, Criterion2, ... to rubric_criteria if possible
            rubric_map = {f"Criterion{i+1}": crit for i, crit in enumerate(rubric_criteria)}

        # Flatten results for DataFrame
        flattened_data = []
        for result in results:
            row = {
                'Student Name': result.get('student_name', ''),
                'Repository URL': result.get('repo_url', ''),
                'AI Percentage': result.get('ai_percentage', 0),
                'Status': result.get('status', 'unknown')
            }
            scores = result.get('scores', {})
            verdict_parts = []
            for idx, criterion in enumerate(rubric_criteria):
                # Try to match rubric criterion to scores key (by title)
                crit_title = criterion.split('(')[0].strip()
                # Find matching key in scores (case-insensitive, ignore extra spaces)
                match_key = next((k for k in scores if k.lower().replace(" ", "") == crit_title.lower().replace(" ", "")), None)
                # If not found, try mapping Criterion1, Criterion2, ... to rubric_criteria
                if not match_key and f"Criterion{idx+1}" in scores:
                    match_key = f"Criterion{idx+1}"
                score_data = scores.get(match_key, {}) if match_key else {}
                if isinstance(score_data, dict):
                    row[criterion] = score_data.get('mark', '')
                    if score_data.get('justification'):
                        verdict_parts.append(f"{criterion}: {score_data['justification']}")
                elif isinstance(score_data, (int, float, str)):
                    row[criterion] = score_data
                else:
                    row[criterion] = ''
            # Add verdict column at the end
            row['Verdict'] = " | ".join(verdict_parts)
            flattened_data.append(row)

        df = pd.DataFrame(flattened_data)

        # Create Excel file
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, sheet_name='Assessment Results', index=False)

        # Format Excel
        workbook = writer.book
        worksheet = writer.sheets['Assessment Results']

        # Header formatting
        for col in range(len(df.columns)):
            cell = worksheet.cell(row=1, column=col + 1)
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            # Enable wrap text for Verdict column
            if df.columns[col] == "Verdict":
                for row in worksheet.iter_rows(min_row=2, min_col=col+1, max_col=col+1):
                    for verdict_cell in row:
                        verdict_cell.alignment = verdict_cell.alignment.copy(wrapText=True)

        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        writer.close()
        output.seek(0)
        return output
    
    async def _generate_summary(self, results: List[Dict]) -> Dict[str, Any]:
        total_students = len(results)
        successful_assessments = sum(1 for r in results if r.get('status') == 'completed')
        avg_ai_percentage = sum(r.get('ai_percentage', 0) for r in results) / total_students if total_students > 0 else 0
        
        return {
            'total_students': total_students,
            'successful_assessments': successful_assessments,
            'failed_assessments': total_students - successful_assessments,
            'average_ai_percentage': round(avg_ai_percentage, 2),
            'success_rate': round((successful_assessments / total_students) * 100, 2) if total_students > 0 else 0
        }
