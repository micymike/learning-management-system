#!/usr/bin/env python3
"""
Test script to verify Excel file handling
"""
import pandas as pd
import os
from io import BytesIO
from openpyxl import load_workbook
from rubric_handler import upload_rubric_file, parse_rubric_lines

def test_excel_parsing():
    """Test Excel file parsing with sample rubric"""
    rubric_path = "sample_rubric.xlsx"
    
    if not os.path.exists(rubric_path):
        print(f"Error: {rubric_path} not found")
        return
    
    print(f"Testing Excel file: {rubric_path}")
    
    # Read the Excel file
    with open(rubric_path, 'rb') as f:
        content = f.read()
    
    print(f"File size: {len(content)} bytes")
    
    # Try to parse with openpyxl
    try:
        print("\nTrying to parse with openpyxl...")
        bytes_io = BytesIO(content)
        workbook = load_workbook(bytes_io)
        print(f"Excel file loaded with openpyxl. Sheets: {workbook.sheetnames}")
        
        # Print sheet contents
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {sheet.dimensions}")
            
            # Print first 5 rows
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i < 5:
                    print(f"Row {i+1}: {row}")
                else:
                    break
    except Exception as e:
        print(f"openpyxl parsing failed: {str(e)}")
    
    # Try to parse with pandas
    try:
        print("\nTrying to parse with pandas...")
        bytes_io = BytesIO(content)
        df = pd.read_excel(bytes_io)
        print(f"Excel file loaded with pandas. Shape: {df.shape}")
        print(f"Columns: {df.columns.tolist()}")
        print("\nFirst 5 rows:")
        print(df.head())
    except Exception as e:
        print(f"pandas parsing failed: {str(e)}")
    
    # Try with rubric_handler
    try:
        print("\nTrying to parse with rubric_handler...")
        rubric_items = upload_rubric_file(content)
        print("\nParsed rubric items:")
        for item in rubric_items:
            print(f"- {item['criterion']}: {item['max_points']} points")
    except Exception as e:
        print(f"rubric_handler parsing failed: {str(e)}")

if __name__ == "__main__":
    test_excel_parsing()